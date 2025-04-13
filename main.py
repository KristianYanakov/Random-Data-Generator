import openpyxl as xl
from openpyxl import Workbook

import generators as gen
from generators import birthYear

#Generate Name and Age
numRows = 10
nameCell = gen.generateNames(numRows)
ageCell = gen.generateAge(numRows)
emailCell = gen.generateEmail(numRows, nameCell)
phoneCell = gen.generatePhoneNumber(numRows)
cityCell = gen.generateCity(numRows)
countryCell = gen.generateCounty(numRows)
birthCell = gen.birthYear(numRows, ageCell)


# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "People Data"

# Write headers
ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['C1'] = 'Birth Year'
ws['D1'] = 'Email'
ws['E1'] = 'Phone Number'
ws['F1'] = 'Country'
ws['G1'] = 'City'


# Write data to cells
for i in range(numRows):
    ws.cell(row = i + 2, column = 1, value = nameCell[i])  # Column A - Name
    ws.cell(row = i + 2, column = 2, value = ageCell[i])   # Column B - Age
    ws.cell(row = i + 2, column = 3, value = birthCell[i])  # Column C - Birth Year
    ws.cell(row = i + 2, column = 4, value = emailCell[i])  # Column D - Email
    ws.cell(row = i + 2, column = 5, value = phoneCell[i])  # Column E - Phone Number
    ws.cell(row = i + 2, column = 6, value = countryCell[i])  # Column F - Country
    ws.cell(row = i + 2, column = 7, value = cityCell[i])  # Column G - City


# Save the workbook
wb.save("fake_data.xlsx")